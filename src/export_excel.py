"""Export a permeability run to a native .xlsx (mirrors the lab spreadsheet).

Builds a real Excel workbook with:
  * the (pressure, flow rate) data table (all replicate points)
  * a native scatter chart with a linear trendline that shows the equation and
    R² — Excel recomputes these itself, so the chart stays editable
  * the Darcy permeability (slope method) and mean hydraulic pore size cells

openpyxl is optional: if it isn't installed, the CSV/JSON/PNG still export and
this is skipped.
"""
from __future__ import annotations

from typing import Optional

SCI = "0.000E+00"

# Provenance columns appended (in this order) to the "Per point" sheet when the
# caller stamps them on the rows. They go at the END: old sheets and any script
# reading by position keep working. Absent from every row -> not written at all,
# so a caller that doesn't stamp them gets exactly the previous workbook.
PROVENANCE_COLS = ["experiment_id", "experiment_label", "ceiling_raised",
                   "collected_under_raised_ceiling"]

RAISED_FILL = "FFF2CC"      # soft amber: this row's run ran above its declared ceiling


def _split_raised(result, points_detail):
    """Partition the raised-ceiling rows into (excluded, kept) with respect to
    THIS fit. Which one applies depends on the caller, and the honest answer is
    whatever actually happened rather than which code path we assume:

      * playlist fit -> Playlist.collected_points() drops raised runs, so those
        rows are absent from result.points_kpa_m3s  -> excluded.
      * single-run fit -> nothing is dropped (it never consults the queue), so
        the rows ARE in the fit -> kept, and the k being reported is derived
        from a run that exceeded its declared envelope. That has to be loud.

    Membership is an exact float compare on purpose: both the fit points and
    these rows are the same values copied from the same TestResult, never
    recomputed. If a future change rounds one side, this silently reports
    everything as excluded -- match it back up rather than loosening to a
    tolerance, so the sheet can't quietly claim a point was dropped.
    """
    rows = [d for d in (points_detail or []) if d.get("ceiling_raised")]
    # only rows that could contribute a point at all
    rows = [d for d in rows if d.get("success") and (d.get("flow_m3s") or 0.0) > 0]
    fit_pts = {(p, q) for p, q in result.points_kpa_m3s}
    excluded, kept = [], []
    for d in rows:
        pt = (d.get("mean_kpa"), d.get("flow_m3s"))
        (kept if pt in fit_pts else excluded).append(d)
    return excluded, kept


def _run_count(rows) -> int:
    """How many distinct runs these rows came from (blank id = one unnamed run)."""
    return len({d.get("experiment_id") or "" for d in rows}) if rows else 0


def xlsx_available() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except Exception:
        return False


def export_permeability_xlsx(result, out_path, *, title: str = "Q vs ΔP",
                             units: str = "kPa", points_detail=None) -> Optional[str]:
    if result.n < 1 or not out_path:
        # see plot_permeability: str(None) would write a workbook named "None"
        return None
    from openpyxl import Workbook
    from openpyxl.chart import Reference, ScatterChart, Series
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.trendline import Trendline
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Permeability"
    bold = Font(bold=True)

    # --- membrane header ------------------------------------------------------
    meta = [
        ("membrane", result.label or "—"),
        ("area (cm²)", result.area_m2 * 1e4),
        ("thickness (mm)", result.thickness_m * 1e3),
        ("water temp (°C)", result.water_temp_c),
        ("viscosity (Pa·s)", result.viscosity_pa_s),
    ]
    for i, (k, v) in enumerate(meta, start=1):
        ws.cell(row=i, column=1, value=k).font = bold
        ws.cell(row=i, column=2, value=v)

    # --- data table -----------------------------------------------------------
    head_row = 6
    ws.cell(row=head_row, column=2, value=f"pressure ({units})").font = bold
    ws.cell(row=head_row, column=3, value="flow rate (m³/s)").font = bold
    first = head_row + 1
    for j, (p, q) in enumerate(result.points_kpa_m3s):
        ws.cell(row=first + j, column=2, value=round(p, 4))
        c = ws.cell(row=first + j, column=3, value=q)
        c.number_format = SCI
    last = first + result.n - 1

    # --- native scatter chart + linear trendline ------------------------------
    chart = ScatterChart()
    chart.title = f"{title} — {result.label}" if result.label else title
    chart.x_axis.title = f"ΔP ({units})"
    chart.y_axis.title = "flow rate  Q  (m³/s)"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.scaling.min = 0
    chart.y_axis.scaling.min = 0
    chart.height = 9
    chart.width = 16
    xref = Reference(ws, min_col=2, min_row=first, max_row=last)
    yref = Reference(ws, min_col=3, min_row=first, max_row=last)
    series = Series(yref, xref, title="Q")
    series.marker = Marker(symbol="circle", size=6)
    series.graphicalProperties.line.noFill = True  # markers only (scatter)
    series.trendline = Trendline(trendlineType="linear", dispEq=True, dispRSqr=True)
    chart.series.append(series)
    ws.add_chart(chart, "E6")

    # --- results block --------------------------------------------------------
    r = last + 2
    verdict = "follows Darcy's law" if result.follows_darcy else "low R² — check linearity"
    rows = [
        ("slope (m³/s per kPa)", result.slope_per_kpa, SCI),
        ("intercept (m³/s)", result.intercept_m3s, SCI),
        ("R²", round(result.r2, 6), None),
        ("Darcy permeability using the slope (m²)", result.k_darcy_m2, SCI),
        ("Mean hydraulic pore size (m)", result.pore_size_m, SCI),
        ("Mean hydraulic pore size (µm)", round(result.pore_size_m * 1e6, 4), None),
        ("verdict", verdict, None),
    ]
    # How well the slope is pinned down, which R² alone does not say. Omitted
    # rather than shown as zero when n < 3, where it is unknown, not zero.
    se = getattr(result, "k_stderr_m2", 0.0)
    if se > 0:
        # The relative error is only meaningful against a physically meaningful
        # k. A negative slope (a leak that opens with pressure, a torn mesh)
        # gives k < 0, and se/k would render a NEGATIVE standard error — the
        # absolute one still describes the scatter, the percentage does not.
        # Same k > 0 guard pore_size_m already uses.
        rows[4:4] = [
            ("k standard error (m²)", se, SCI),
            ("k standard error (%)", round(se / result.k_darcy_m2 * 100, 2)
             if result.k_darcy_m2 > 0 else None, None),
        ]
    for i, (label, val, fmt) in enumerate(rows):
        ws.cell(row=r + i, column=1, value=label).font = bold
        cell = ws.cell(row=r + i, column=2, value=val)
        if fmt:
            cell.number_format = fmt

    # --- ceiling-raised provenance -------------------------------------------
    # A k derived from (or missing points because of) a run that raised its
    # ceiling must be visible right here next to the number, not only in the
    # raw CSV. Silent on a clean dataset: no raised rows -> no lines added.
    excluded, kept = _split_raised(result, points_detail)
    notes = []
    if excluded:
        notes.append((
            "excluded from this fit",
            f"{len(excluded)} point(s) from {_run_count(excluded)} run(s) — ceiling raised mid-run"))
    if kept:
        notes.append((
            "⚠ ceiling raised — in this fit",
            f"{len(kept)} point(s) of this k came from a run whose ceiling was raised"))
    nr = r + len(rows) + 1
    for i, (label, text) in enumerate(notes):
        ws.cell(row=nr + i, column=1, value=label).font = bold
        ws.cell(row=nr + i, column=2, value=text)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16

    # --- optional per-point summary sheet ------------------------------------
    if points_detail:
        from openpyxl.styles import PatternFill

        ps = wb.create_sheet("Per point")
        cols = ["setpoint_kpa", "mean_kpa", "std_kpa", "min_kpa", "max_kpa",
                "in_band_fraction", "n_samples", "collection_s", "volume_ml", "flow_m3s"]
        # This sheet mixes rows from several runs, so provenance has to be
        # per-row; a scalar couldn't label it. Only append the columns the
        # caller actually stamped.
        cols += [c for c in PROVENANCE_COLS if any(c in d for d in points_detail)]
        for c, name in enumerate(cols, start=1):
            ps.cell(row=1, column=c, value=name).font = bold
        amber = PatternFill("solid", fgColor=RAISED_FILL)
        for ri, d in enumerate(points_detail, start=2):
            for c, name in enumerate(cols, start=1):
                cell = ps.cell(row=ri, column=c, value=d.get(name))
                if name == "flow_m3s":
                    cell.number_format = SCI
                if d.get("ceiling_raised"):
                    cell.fill = amber      # findable by eye, not just by column

    out_path = str(out_path)
    wb.save(out_path)
    return out_path
