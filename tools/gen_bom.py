"""Generate the purchase BOM (BOM.csv, BOM.xlsx, and the Desktop copy).

Single source of truth for the shopping list — edit ROWS and re-run:
    python tools/gen_bom.py
Rows in a group starting "OPTIONAL" are excluded from the required total and
shaded yellow; "Fittings" rows are shaded green (measure tube OD first).
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

DESKTOP_COPY = "/Users/salvador/Desktop/BOM_MembraneRig.xlsx"

# (group, part, qty, unit_usd, supplier, url, notes)
ROWS = [
 ("Control (air line, on existing valve)", "Metal-gear servo DS3218 (20 kg·cm, 270°)", 1, 18.00, "Amazon", "https://www.amazon.com/ANNIMOS-Digital-Waterproof-DS3218MG-Control/dp/B076CNKQX4", "Turns the EXISTING SS air ball valve (quarter-turn)"),
 ("Control (air line, on existing valve)", "Servo↔ball-valve coupling + servo mount", 1, 0.00, "In-house 3D print", "", "Printed by you, onto the existing green ball valve's stem (handle removed)"),
 ("Control (air line, on existing valve)", "UBEC 12V→6V 3A (servo power)", 1, 12.00, "Amazon", "https://www.amazon.com/Hobbywing-UBEC-3A-Input-Switch-Mode-Regulator/dp/B07T2CKC8G", "Powers servo off 12V, NOT the Pi"),
 ("Sensing & electronics", "Official Raspberry Pi 15W USB-C PSU (SC0218)", 1, 10.00, "Amazon", "https://www.amazon.com/Raspberry-Model-Official-SC0218-Accessory/dp/B07W8XHMJZ", "SKIP if you already have it"),
 ("Sensing & electronics", "Pressure transducer 0–15 PSI, 0.5–4.5V, G1/4", 1, 15.00, "Amazon", "https://www.amazon.com/0-5-4-5V-Stainless-Pressure-Transducer-0-15PSI/dp/B07G5DQZJ2", "0-15 PSI = 0-103 kPa; tests top ~60 kPa"),
 ("Sensing & electronics", "ADS1115 16-bit I2C ADC", 1, 9.00, "Amazon", "https://www.amazon.com/HiLetgo-Converter-Programmable-Amplifier-Development/dp/B01DLHKMO2", "Power at 3.3V; sensor via 10k/20k divider"),
 ("Sensing & electronics", "DS18B20 waterproof temperature probe (1-Wire)", 1, 8.00, "Amazon", "https://www.amazon.com/s?k=DS18B20+waterproof+temperature+sensor+probe+stainless", "NEW: measures water temp -> mu auto-computed. Data->GPIO4 + 4.7k pullup (from kit)"),
 ("Sensing & electronics", "Resistor assortment kit (1% metal film)", 1, 14.00, "Amazon", "https://www.amazon.com/Resistor-Assorted-Resistors-Assortment-Experiments/dp/B07L851T3V", "Divider + MOSFET gate/pulldown + DS18B20 4.7k pullup"),
 ("Sensing & electronics", "IRLZ44N logic-level MOSFET (10-pack)", 1, 9.00, "Amazon", "https://www.amazon.com/YINETTECH-Transistors-29x10x4-5mm-Electronic-Controllable/dp/B0D3Q24JH6", "Drives the 3-way diverter; add 10k gate pulldown"),
 ("Sensing & electronics", "1N5819 Schottky diode (100-pack)", 1, 8.00, "Amazon", "https://www.amazon.com/100-Pieces-1N5819-Schottky-Rectifier/dp/B079KG1TN2", "Flyback across the solenoid"),
 ("Sensing & electronics", "12V 3A DC power supply (5.5×2.1mm)", 1, 11.00, "Amazon", "https://www.amazon.com/ANVISION-Supply-Barrel-5-5x2-1mm-Efficiency/dp/B01C010YJI", "Feeds solenoid + UBEC"),
 ("Sensing & electronics", "Adjustable relief valve CR25-100 (SAFETY)", 1, 18.00, "Amazon", "https://www.amazon.com/Control-Devices-Pressure-Relief-Adjustable/dp/B007GDY3CU", "On the air side; SKIP if the gas panel already relieves"),
 ("Fluidics & proto", "Solderless breadboard 830 + jumper kit", 1, 12.00, "Amazon", "https://www.amazon.com/BOJACK-Values-Solderless-Breadboard-Flexible/dp/B08Y59P6D1", ""),
 ("Fluidics & proto", "Silicone tubing 1/4\" ID (permeate side)", 1, 15.00, "Amazon", "https://www.amazon.com/Silicone-Tubing-JoyTube-Black-Transfer/dp/B0BLZ3PB9Q", "Low-pressure permeate line only"),
 ("Fluidics & proto", "Hose barb fitting kit (barb-to-barb)", 1, 12.00, "Amazon", "https://www.amazon.com/Fittings-Assortment-Connector-Catheter-Adapters/dp/B0CSKH6JXR", "Permeate tubing joins"),
 ("Fittings — MEASURE tube OD first (Swagelok rig)", "Transducer adapter: G1/4-M → Swagelok (size = tube OD)", 1, 10.00, "McMaster/Swagelok", "https://www.mcmaster.com/products/tube-fittings/", "Rig is compression/Swagelok, NOT NPT. Or mount at the existing manometer port"),
 ("Fittings — MEASURE tube OD first (Swagelok rig)", "Swagelok tee for sensor/relief tap (size = tube OD)", 1, 12.00, "McMaster/Swagelok", "https://www.mcmaster.com/products/tube-fittings/", "CHECK LAB STOCK first"),
 ("Fittings — MEASURE tube OD first (Swagelok rig)", "1/4\" NPT-M × 1/4\" barb adapter", 3, 1.40, "Amazon", "https://www.amazon.com/s?k=1%2F4+NPT+male+to+1%2F4+hose+barb+brass", "For the 3-way diverter ports -> silicone (permeate side)"),
 ("Fittings — MEASURE tube OD first (Swagelok rig)", "PTFE thread-seal tape (roll)", 1, 2.00, "Amazon", "https://www.amazon.com/s?k=PTFE+thread+seal+tape", "NPT (diverter) joints only"),
 ("Fittings — MEASURE tube OD first (Swagelok rig)", "Hose clamps (worm/Oetiker, ~10)", 1, 7.00, "Amazon", "https://www.amazon.com/s?k=1%2F4+inch+hose+clamp+small+assortment", "Every barb-to-silicone joint"),
 ("Fittings — MEASURE tube OD first (Swagelok rig)", "Barrel jack 5.5×2.1mm → screw terminal", 1, 6.00, "Amazon", "https://www.amazon.com/s?k=5.5x2.1mm+female+barrel+jack+to+screw+terminal+adapter", "Land the 12V into the breadboard"),
 ("Fittings — MEASURE tube OD first (Swagelok rig)", "Inline blade-fuse holder + 3A fuse", 1, 6.00, "Amazon", "https://www.amazon.com/s?k=inline+blade+fuse+holder+12v+3A", "Fire safety on the 12V"),
 ("Fittings — MEASURE tube OD first (Swagelok rig)", "M3 screws/standoffs + baseplate", 1, 10.00, "Amazon", "https://www.amazon.com/s?k=M3+standoff+screw+assortment+kit", "Mount Pi/ADS/servo bracket"),
 ("Fittings — MEASURE tube OD first (Swagelok rig)", "22AWG stranded hookup wire (spool)", 1, 8.00, "Amazon", "https://www.amazon.com/s?k=22AWG+stranded+hookup+wire+kit", "Reach solenoid/servo/12V/probe"),
 ("Fittings — MEASURE tube OD first (Swagelok rig)", "Electrolytic caps 1000µF + 470µF", 1, 3.00, "Amazon", "https://www.amazon.com/s?k=1000uf+470uf+25v+electrolytic+capacitor", "Rail stability"),
 ("Fittings — MEASURE tube OD first (Swagelok rig)", "Pi/electronics enclosure (splash protection)", 1, 0.00, "In-house 3D print", "", "Designed & printed by you"),
 ("Fittings — MEASURE tube OD first (Swagelok rig)", "Zip ties + heat-shrink assortment", 1, 8.00, "Amazon", "https://www.amazon.com/s?k=heat+shrink+tubing+zip+ties+assortment+kit", "Strain relief + insulate splices"),
 ("Order separately (not Amazon)", "3-way 12V solenoid valve for WATER (diverter)", 1, 53.00, "ESValves", "https://www.electricsolenoidvalves.com/1-4-3-way-12v-dc-electric-solenoid-valve/", "On the permeate outlet; Amazon 3-way are air/gas only"),
 ("OPTIONAL — feed-forward upgrade", "2nd pressure transducer 0–15 PSI (supply side)", 1, 15.00, "Amazon", "https://www.amazon.com/0-5-4-5V-Stainless-Pressure-Transducer-0-15PSI/dp/B07G5DQZJ2", "Only if compressor swings beat the PID"),
 ("OPTIONAL — feed-forward upgrade", "G1/4-M → Swagelok adapter (2nd sensor)", 1, 10.00, "McMaster/Swagelok", "https://www.mcmaster.com/products/tube-fittings/", "Mount for the supply-side sensor"),
]

NOTES = [
 "RIG = air-over-water: compressed AIR pressurises the vessel. The servo turns the EXISTING stainless air BALL VALVE (quarter-turn; no new control valve).",
 "TEMPERATURE is a test variable (distilled water). The DS18B20 probe measures the water temp -> the software computes mu(T) and logs it. k comes out ~constant across temperature (validation).",
 "Plumbing is Swagelok/COMPRESSION, not NPT. MEASURE the tube OD (caliper) before buying green-shaded fittings; many are likely in lab stock. Mount the transducer at the existing manometer port if possible.",
 "Sensor 0-15 PSI (0-103 kPa) for <=60 kPa tests. ADS1115 at 3.3V + 10k/20k divider. DS18B20 on GPIO4 + 4.7k pullup.",
 "Control is COARSE (servo trims a quarter-turn ball valve). k is still valid: Q is regressed vs the MEASURED mean pressure per point.",
 "OPTIONAL (yellow): supply-side sensor for feed-forward, only if needed.",
 "FIRST TEST once assembled: sweep the servo across the ball valve's travel and log pressure BEFORE tuning PID; use the dial gauge for a 2-point sensor cal.",
]


def totals():
    req = sum(q * u for g, _, q, u, _, _, _ in ROWS if not g.startswith("OPTIONAL"))
    opt = sum(q * u for g, _, q, u, _, _, _ in ROWS if g.startswith("OPTIONAL"))
    return req, opt


def write_csv(path):
    req, opt = totals()
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Group", "Part", "Qty", "Unit (USD)", "Subtotal (USD)", "Supplier", "Link", "Notes"])
        for g, p, q, u, s, url, n in ROWS:
            w.writerow([g, p, q, f"{u:.2f}", f"{q*u:.2f}", s, url, n])
        w.writerow([])
        w.writerow(["", "TOTAL required (USD)", "", "", f"{req:.2f}", "", "", "Swagelok items may be ~$0 from lab stock"])
        w.writerow(["", "TOTAL with optional (USD)", "", "", f"{req+opt:.2f}", "", "", "Already owned: Pi 4, microSD, graduated cylinder, dial gauge, air ball valve."])


def write_xlsx(path):
    req, opt = totals()
    wb = Workbook(); ws = wb.active; ws.title = "BOM"
    hdr = ["Group", "Part", "Qty", "Unit (USD)", "Subtotal (USD)", "Supplier", "Link", "Notes"]
    fill = PatternFill("solid", fgColor="2F5597")
    border = Border(bottom=Side(style="thin", color="DDDDDD"))
    optfill = PatternFill("solid", fgColor="FFF2CC")
    swfill = PatternFill("solid", fgColor="E2EFDA")
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(1, c, h); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = fill
    r = 2
    for g, p, q, u, s, url, n in ROWS:
        ws.cell(r, 1, g); ws.cell(r, 2, p); ws.cell(r, 3, q)
        ws.cell(r, 4, u).number_format = '$#,##0.00'
        ws.cell(r, 5, q * u).number_format = '$#,##0.00'
        ws.cell(r, 6, s)
        lc = ws.cell(r, 7, "link" if url else "")
        if url:
            lc.hyperlink = url; lc.font = Font(color="0563C1", underline="single")
        ws.cell(r, 8, n)
        for c in range(1, 9):
            ws.cell(r, c).border = border
            if g.startswith("OPTIONAL"):
                ws.cell(r, c).fill = optfill
            elif g.startswith("Fittings"):
                ws.cell(r, c).fill = swfill
        r += 1
    ws.cell(r + 1, 2, "TOTAL required (USD)").font = Font(bold=True)
    tv = ws.cell(r + 1, 5, req); tv.font = Font(bold=True); tv.number_format = '$#,##0.00'
    ws.cell(r + 2, 2, "TOTAL with optional (USD)").font = Font(bold=True)
    tv2 = ws.cell(r + 2, 5, req + opt); tv2.font = Font(bold=True); tv2.number_format = '$#,##0.00'
    for i, t in enumerate(NOTES):
        ws.cell(r + 4 + i, 2, t).font = Font(italic=True, size=10)
    for c, wd in enumerate([34, 52, 6, 12, 14, 18, 10, 58], 1):
        ws.column_dimensions[chr(64 + c)].width = wd
    ws.freeze_panes = "A2"
    wb.save(path)


if __name__ == "__main__":
    write_csv("BOM.csv")
    write_xlsx("BOM.xlsx")
    try:
        write_xlsx(DESKTOP_COPY)
    except Exception as e:
        print("desktop copy skipped:", e)
    req, opt = totals()
    print(f"{len(ROWS)} items | required ${req:.2f} | with optional ${req+opt:.2f}")
